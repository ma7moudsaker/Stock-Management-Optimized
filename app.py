import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime
import json
from database import StockDatabase
from io import BytesIO
# ✅ openpyxl imports
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
import atexit
import threading
import time
from dropbox_oauth_backup import DropboxOAuthBackup
# أضف هذا مع الـ imports في الأعلى
from barcode_utils import (
    BarcodeGenerator, 
    BarcodePrinter,
    generate_barcode_for_variant,
    create_barcode_labels_pdf,
    validate_ean13
)

app = Flask(__name__)

# إعدادات الأمان والإنتاج
app.secret_key = os.environ.get('SECRET_KEY', 'fallback-secret-for-dev')

# Session timeout (30 يوم)
app.config['PERMANENT_SESSION_LIFETIME'] = 2592000  # 30 days in seconds

# Remove or comment out the old simple login_required
# We'll create a better one

def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Please login to continue', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def page_permission_required(page_key):
    """Decorator to check if user has permission for a specific page"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('logged_in'):
                flash('Please login to continue', 'warning')
                return redirect(url_for('login'))
            
            user_id = session.get('user_id')
            
            # Super Admin (from env) has all permissions
            if user_id == 0:
                return f(*args, **kwargs)
            
            # Check database permission
            if not db.user_has_permission(user_id, page_key):
                flash('You do not have permission to access this page!', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def action_permission_required(page_key):
    """Decorator for action routes that inherit permission from parent page"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('logged_in'):
                flash('Please login to continue', 'warning')
                return redirect(url_for('login'))
            
            user_id = session.get('user_id')
            
            # Super Admin has all permissions
            if user_id == 0:
                return f(*args, **kwargs)
            
            # Check permission
            if not db.user_has_permission(user_id, page_key):
                flash('You do not have permission to perform this action!', 'error')
                # Return to referrer or dashboard
                return redirect(request.referrer or url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


@app.context_processor
def inject_user_data():
    """Inject user permissions and info into all templates"""
    user_id = session.get('user_id', 0)
    
    if user_id:
        # Temporary: Give all permissions for testing
        user_permissions = {
            'dashboard': True,
            'products_new': True,
            'add_product_new': True,
            'add_products_multi': True,
            'barcode_system': True,
            'inventory_management': True,
            'bulk_upload_excel': True,
            'export_products': True,
            'manage_brands': True,
            'manage_colors': True,
            'manage_product_types': True,
            'manage_tags': True,
            'manage_trader_categories': True,
            'user_management': True,
            'logs': True,
            'backup_system': True
        }
        
        return {
            'user_permissions': user_permissions,
            'user_id': user_id,
            'username': session.get('username', ''),
            'full_name': session.get('full_name', ''),
            'is_super_admin': True
        }
    
    return {
        'user_permissions': {},
        'user_id': 0,
        'username': '',
        'full_name': '',
        'is_super_admin': False
    }

@app.route('/user_management')
@page_permission_required('user_management')
def user_management():
    """User Management Page"""
    users = db.get_users_with_permissions()
    pages_grouped = db.get_all_pages()
    
    # Count total pages
    total_pages = sum(len(pages) for pages in pages_grouped.values())
    
    return render_template('user_management.html', 
                         users=users,
                         pages_grouped=pages_grouped,
                         total_pages=total_pages)

@app.route('/add_user', methods=['POST'])
@page_permission_required('user_management')
def add_user():
    """Add a new user"""
    try:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        full_name = request.form.get('full_name', '').strip()
        page_keys = request.form.getlist('page_keys')
        
        if not username or not password or not full_name:
            flash('Username, password, and full name are required!', 'error')
            return redirect(url_for('user_management'))
        
        # Check if username already exists
        existing_user = db.get_user_by_username(username)
        if existing_user:
            flash(f'Username "{username}" already exists!', 'error')
            return redirect(url_for('user_management'))
        
        # Create user
        granted_by = session.get('user_id', 0)
        user_id = db.create_user(username, password, full_name, granted_by)
        
        if user_id:
            # Set permissions
            if page_keys:
                db.set_user_permissions(user_id, page_keys, granted_by)
            
            flash(f'User "{full_name}" created successfully with {len(page_keys)} permissions! ✅', 'success')
        else:
            flash('Error creating user!', 'error')
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('user_management'))

@app.route('/edit_user/<int:user_id>', methods=['POST'])
@page_permission_required('user_management')
def edit_user(user_id):
    """Edit user information and permissions"""
    try:
        username = request.form.get('username', '').strip()
        full_name = request.form.get('full_name', '').strip()
        password = request.form.get('password', '').strip()
        active = int(request.form.get('active', 1))
        page_keys = request.form.getlist('page_keys')
        
        # Update user info
        db.update_user(
            user_id, 
            username=username if username else None,
            full_name=full_name if full_name else None,
            password=password if password else None,
            active=active
        )
        
        # Update permissions
        granted_by = session.get('user_id', 0)
        db.set_user_permissions(user_id, page_keys, granted_by)
        
        flash(f'User updated successfully! ✅', 'success')
        
    except Exception as e:
        flash(f'Error updating user: {str(e)}', 'error')
    
    return redirect(url_for('user_management'))

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@page_permission_required('user_management')
def delete_user(user_id):
    """Delete a user"""
    try:
        user = db.get_user_by_id(user_id)
        if user:
            db.delete_user(user_id)
            flash(f'User "{user["full_name"]}" deleted successfully! 🗑️', 'success')
        else:
            flash('User not found!', 'error')
    except Exception as e:
        flash(f'Error deleting user: {str(e)}', 'error')
    
    return redirect(url_for('user_management'))

@app.route('/toggle_user/<int:user_id>', methods=['POST'])
@page_permission_required('user_management')
def toggle_user(user_id):
    """Toggle user active status"""
    try:
        user = db.get_user_by_id(user_id)
        if user:
            new_status = 0 if user['active'] == 1 else 1
            db.update_user(user_id, active=new_status)
            status_text = 'activated' if new_status == 1 else 'deactivated'
            flash(f'User "{user["full_name"]}" {status_text} successfully!', 'success')
        else:
            flash('User not found!', 'error')
    except Exception as e:
        flash(f'Error toggling user status: {str(e)}', 'error')
    
    return redirect(url_for('user_management'))

@app.route('/get_user_permissions/<int:user_id>')
@page_permission_required('user_management')
def get_user_permissions(user_id):
    """Get user permissions as JSON"""
    permissions = db.get_user_permissions(user_id)
    return jsonify(permissions)


# إنشاء قاعدة البيانات
print("🔄 Initializing database...")
db = StockDatabase()
print("✅ Database initialized!")

# إضافة البيانات الافتراضية فقط في البيئة المحلية
if not os.getenv('DATABASE_URL'):
    db.add_default_data()
    print("✅ Default data added!")


# إنشاء نظام النسخ الاحتياطية
backup_system = DropboxOAuthBackup()

# متغير للتأكد من تشغيل الكود مرة واحدة فقط
startup_completed = False
@app.before_request
def restore_on_startup():
    global startup_completed
    if not startup_completed:
        try:
            # === Restore Database ===
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM base_products')
            brand_count = cursor.fetchone()[0]
            
            # === Check and Restore Logs ===
            cursor.execute('SELECT COUNT(*) FROM stock_logs')
            logs_count = cursor.fetchone()[0]
            conn.close()
            
            # Restore database if empty
            if brand_count == 0:
                print("⚠️ Database is empty. Attempting restore from Dropbox...")
                success = backup_system.restore_from_backup()
                if not success:
                    print("⚠️ No backup found. Adding default data...")
                    db.add_default_data()
                else:
                    print("✅ Database restored from backup!")
            else:
                print(f"✅ Database OK: {brand_count} products found")
            
            # Restore logs if empty
            if logs_count == 0 and backup_system.dbx:
                print("⚠️ Logs table is empty. Attempting to restore from backup...")
                try:
                    # List all log backups
                    result = backup_system.dbx.files_list_folder('/Stock_Backups')
                    if result.entries:
                        # Get latest backup
                        latest_backup = sorted(result.entries, key=lambda x: x.name, reverse=True)[0]
                        
                        # Download and restore
                        _, response = backup_system.dbx.files_download(f'/Stock_Backups/{latest_backup.name}')
                        with open('temp_logs_restore.json', 'wb') as f:
                            f.write(response.content)
                        
                        db.import_logs_from_json('temp_logs_restore.json')
                        os.remove('temp_logs_restore.json')
                        
                        print(f"✅ Logs restored from: {latest_backup.name}")
                    else:
                        print("⚠️ No logs backup found")
                except Exception as e:
                    print(f"❌ Error restoring logs: {e}")
            else:
                print(f"✅ Logs OK: {logs_count} entries found")
                
        except Exception as e:
            print(f"❌ Error in startup restoration: {e}")
            db.add_default_data()
        
        startup_completed = True
        print("✅ Startup completed!")

def auto_backup():
    """نسخ احتياطية تلقائية كل 6 ساعة"""
    while True:
        time.sleep(21600)  # كل 6 ساعة
        print("🔄 إنشاء نسخة احتياطية تلقائية...")
        backup_system.create_backup()

# === LOGS BACKUP SYSTEM ===
def daily_logs_backup():
    """Backup logs to Dropbox every 24 hours"""
    while True:
        time.sleep(86400)  # 24 hours
        try:
            print("🔄 Starting daily logs backup...")
            
            # Export to JSON
            backup_filename = f'logs_backup_{datetime.now().strftime("%Y-%m-%d")}.json'
            
            if db.export_logs_to_json(backup_filename):
                # Upload to Dropbox
                if backup_system.dbx:
                    try:
                        with open(backup_filename, 'rb') as f:
                            backup_system.dbx.files_upload(
                                f.read(),
                                f'/Stock_Backups/{backup_filename}',
                                mode=dropbox.files.WriteMode.overwrite
                            )
                        print(f"✅ Logs backup uploaded: {backup_filename}")
                        
                        # Delete local file
                        os.remove(backup_filename)
                    except Exception as e:
                        print(f"❌ Error uploading logs backup: {e}")
                        if os.path.exists(backup_filename):
                            os.remove(backup_filename)
            
        except Exception as e:
            print(f"❌ Error in logs backup: {e}")

# Start logs backup thread
logs_backup_thread = threading.Thread(target=daily_logs_backup, daemon=True)
logs_backup_thread.start()
print("✅ Logs backup system started")


# بدء النسخ الاحتياطية التلقائية
backup_thread = threading.Thread(target=auto_backup)
backup_thread.daemon = True
backup_thread.start()
print("✅ تم بدء نظام النسخ الاحتياطية التلقائية (Dropbox)")

# نسخة احتياطية عند إغلاق التطبيق
@atexit.register
def backup_on_exit():
    print("🔄 إنشاء نسخة احتياطية قبل الإغلاق...")
    # إضافة تأخير للتأكد من اكتمال العمليات
    time.sleep(3)
    backup_system.create_backup()

# routes الجديدة للنسخ الاحتياطية
@app.route('/admin/backup')
@page_permission_required('backup_system')
def backup_page():
    """صفحة إدارة النسخ الاحتياطية"""
    backups = backup_system.list_backups()
    return render_template('backup_system.html', backups=backups, service="Dropbox")

@app.route('/admin/backup/create')
@action_permission_required('backup_system')
def create_backup():
    """إنشاء نسخة احتياطية فورية"""
    success = backup_system.create_backup()
    if success:
        flash('تم إنشاء النسخة الاحتياطية في Dropbox بنجاح!', 'success')
    else:
        flash('خطأ في إنشاء النسخة الاحتياطية', 'error')
    
    return redirect(url_for('backup_page'))

@app.route('/admin/backup/status')
@action_permission_required('backup_system')
def backup_status():
    """حالة نظام النسخ الاحتياطية"""
    backups = backup_system.list_backups()
    status = {
        'service': 'Dropbox',
        'connected': backup_system.dbx is not None,
        'backup_count': len(backups),
        'latest_backup': backups[0]['name'] if backups else 'لا توجد نسخ'
    }
    return jsonify(status)

@app.route('/admin/backup/restore/<backup_name>')
@action_permission_required('backup_system')
def restore_backup(backup_name):
    """استرجاع نسخة احتياطية محددة"""
    success = backup_system.restore_from_backup(backup_name)
    if success:
        flash(f'تم استرجاع البيانات من {backup_name} بنجاح!', 'success')
    else:
        flash('فشل في استرجاع البيانات', 'error')
    
    return redirect(url_for('backup_page'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page - Hybrid system (Super Admin + DB Users)"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        # 1️⃣ Check Super Admin
        SUPER_ADMIN_USERNAME = os.environ.get('SUPER_ADMIN_USERNAME', 'admin')
        SUPER_ADMIN_PASSWORD = os.environ.get('SUPER_ADMIN_PASSWORD', 'admin123456')
        
        if username == SUPER_ADMIN_USERNAME and password == SUPER_ADMIN_PASSWORD:
            session['logged_in'] = True
            session['user_id'] = 0
            session['username'] = username
            session['full_name'] = 'Super Admin'
            session['role'] = 'Super Admin'
            flash('Welcome Super Admin! 👑', 'success')
            return redirect(url_for('dashboard'))
        
        # 2️⃣ Check Database Users
        user = db.get_user_by_username(username)
        
        if user and user['active'] == 1:
            if check_password_hash(user['password_hash'], password):
                session['logged_in'] = True
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['full_name'] = user['full_name']
                session['role'] = 'User'
                
                # Update last login
                db.update_last_login(user['id'])
                
                flash(f'Welcome back, {user["full_name"]}! 👋', 'success')
                
                # ✨ Smart redirect: Send to first available page
                redirect_url = get_first_available_page(user['id'])
                return redirect(redirect_url)
            else:
                flash('Invalid password!', 'error')
        else:
            flash('Invalid username or account is disabled!', 'error')
        
        return redirect(url_for('login'))
    
    # GET request
    return render_template('login.html')


def get_first_available_page(user_id):
    """Get the first page the user has permission to access"""
    # Get user's permissions
    permissions = db.get_user_permissions(user_id)
    
    if not permissions:
        # No permissions - show error page or limited dashboard
        flash('You have no page access! Please contact admin.', 'warning')
        return url_for('no_access')
    
    # Priority order for redirect
    priority_pages = [
        ('dashboard', 'dashboard'),
        ('products', 'products_new'),
        ('bulk_inventory', 'inventory_management'),
        ('add_product', 'add_product_new'),
        ('manage_brands', 'manage_brands'),
        ('user_management', 'user_management'),
    ]
    
    # Check priority pages first
    for page_key, route_name in priority_pages:
        if page_key in permissions:
            return url_for(route_name)
    
    # If no priority page found, get first available
    if permissions:
        # Map page_key to route name
        page_route_map = {
            'products': 'products_new',
            'product_details': 'products_new',  # Fallback to products
            'activity_logs': 'logs',
            'add_product': 'add_product_new',
            'add_multiple': 'add_products_multi',
            'bulk_upload': 'bulk_upload_excel',
            'bulk_inventory': 'inventory_management',
            'export_products': 'export_products',
            'edit_product': 'products_new',  # Fallback
            'manage_brands': 'manage_brands',
            'manage_colors': 'manage_colors',
            'manage_product_types': 'manage_product_types',
            'manage_trader_categories': 'manage_trader_categories',
            'manage_tags': 'manage_tags',
            'backup_system': 'backup_page',
            'user_management': 'user_management',
        }
        
        first_page = permissions[0]
        route_name = page_route_map.get(first_page, 'dashboard')
        return url_for(route_name)
    
    # Fallback
    return url_for('no_access')


@app.route('/no_access')
@login_required
def no_access():
    """Page shown when user has no permissions"""
    return render_template('no_access.html')

@app.route('/logout')
def logout():
    """Logout user"""
    username = session.get('full_name', 'User')
    session.clear()
    flash(f'Goodbye {username}! See you soon 👋', 'info')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    """Dashboard - Available to all logged-in users"""
    user_id = session.get('user_id')
    
    # Check if user has ANY permissions
    if user_id != 0:
        user_permissions = db.get_user_permissions(user_id)
        if not user_permissions:
            flash('⚠️ You have no page access! Please contact admin.', 'warning')
            return render_template('no_access.html')
    
    # Check if user is admin (for stock value access)
    is_admin = False
    if user_id == 0:  # Super Admin
        is_admin = True
    elif user_id:
        user_permissions = db.get_user_permissions(user_id)
        if 'user_management' in user_permissions:
            is_admin = True
    
    # Get basic statistics
    products = db.get_all_products_with_details()
    total_products = len(products)
    total_stock_qty = db.get_total_stock_quantity()
    
    # Initialize with default values
    stock_value = 0
    stock_value_trend = {'dates': [], 'values': []}
    
    # Get stock value (only for admins)
    if is_admin:
        try:
            stock_value = db.get_total_stock_value()
            stock_value_trend = db.get_stock_value_trend(days=30)
            
            # ✅ Ensure it's a dict with proper structure
            if not isinstance(stock_value_trend, dict):
                stock_value_trend = {'dates': [], 'values': []}
            if 'dates' not in stock_value_trend:
                stock_value_trend['dates'] = []
            if 'values' not in stock_value_trend:
                stock_value_trend['values'] = []
        except Exception as e:
            print(f"Error getting stock value data: {e}")
            stock_value = 0
            stock_value_trend = {'dates': [], 'values': []}
    
    # Get trends and analytics
    try:
        stock_qty_trend = db.get_stock_quantity_trend(days=30)
        
        # ✅ Ensure proper structure
        if not isinstance(stock_qty_trend, dict):
            stock_qty_trend = {'dates': [], 'quantities': []}
        if 'dates' not in stock_qty_trend:
            stock_qty_trend['dates'] = []
        if 'quantities' not in stock_qty_trend:
            stock_qty_trend['quantities'] = []
            
    except Exception as e:
        print(f"Error getting stock quantity trend: {e}")
        stock_qty_trend = {'dates': [], 'quantities': []}
    
    try:
        most_updated = db.get_most_updated_products(limit=10, days=30)
    except Exception as e:
        print(f"Error getting most updated products: {e}")
        most_updated = []
    
    try:
        top_brands = db.get_top_brands(limit=5)
    except Exception as e:
        print(f"Error getting top brands: {e}")
        top_brands = []

    try:
        top_products_stock = db.get_top_products_by_stock(limit=10)
    except Exception as e:
        print(f"Error getting top products by stock: {e}")
        top_products_stock = []

    
    try:
        products_by_category = db.get_products_by_category()
    except Exception as e:
        print(f"Error getting products by category: {e}")
        products_by_category = []
    
    try:
        system_counts = db.get_active_system_counts()
    except Exception as e:
        print(f"Error getting system counts: {e}")
        system_counts = {
            'brands': 0,
            'categories': 0,
            'types': 0,
            'colors': 0
        }
    
    # Prepare data for template
    stats = {
        'total_products': total_products,
        'total_stock_qty': total_stock_qty,
        'stock_value': stock_value
    }
    
    return render_template('dashboard.html',
                         stats=stats,
                         is_admin=is_admin,
                         stock_qty_trend=stock_qty_trend,
                         stock_value_trend=stock_value_trend,
                         most_updated=most_updated,
                         top_brands=top_brands,
                         top_products_stock=top_products_stock,  # ✅ أضف ده
                         products_by_category=products_by_category,
                         system_counts=system_counts)

# صفحات إدارة البراندات
@app.route('/manage_brands')
@page_permission_required('manage_brands')
def manage_brands():
    brands = db.get_all_brands()
    return render_template('manage_brands.html', brands=brands)

@app.route('/add_brand', methods=['POST'])
@action_permission_required('manage_brands')
def add_brand():
    brand_name = request.form['brand_name'].strip()
    if brand_name:
        if db.add_brand(brand_name):
            flash(f'Brand "{brand_name}" added successfully!', 'success')
        else:
            flash(f'Error: Brand "{brand_name}" already exists!', 'error')
    return redirect(url_for('manage_brands'))

@app.route('/edit_brand/<int:brand_id>', methods=['POST'])
@action_permission_required('manage_brands')
def edit_brand(brand_id):
    new_name = request.form['brand_name'].strip()
    if new_name:
        if db.update_brand(brand_id, new_name):
            flash(f'Brand updated to "{new_name}" successfully!', 'success')
        else:
            flash('Error updating brand - name might already exist!', 'error')
    return redirect(url_for('manage_brands'))

@app.route('/delete_brand/<int:brand_id>', methods=['POST'])
@action_permission_required('manage_brands')
def delete_brand(brand_id):
    success, message = db.delete_brand(brand_id)
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    return redirect(url_for('manage_brands'))

# صفحات إدارة الألوان
@app.route('/manage_colors')
@page_permission_required('manage_colors')
def manage_colors():
    colors = db.get_all_colors()
    return render_template('manage_colors.html', colors=colors)

@app.route('/add_color', methods=['POST'])
@action_permission_required('manage_colors')
def add_color():
    color_name = request.form['color_name'].strip()
    color_code = request.form['color_code']
    if color_name:
        if db.add_color(color_name, color_code):
            flash(f'Color "{color_name}" added successfully!', 'success')
        else:
            flash(f'Error: Color "{color_name}" already exists!', 'error')
    return redirect(url_for('manage_colors'))

@app.route('/edit_color/<int:color_id>', methods=['POST'])
@action_permission_required('manage_colors')
def edit_color(color_id):
    new_name = request.form['color_name'].strip()
    new_code = request.form['color_code'].strip()
    if new_name and new_code:
        if db.update_color(color_id, new_name, new_code):
            flash(f'Color updated to "{new_name}" successfully!', 'success')
        else:
            flash('Error updating color - name might already exist!', 'error')
    return redirect(url_for('manage_colors'))

@app.route('/delete_color/<int:color_id>', methods=['POST'])
@action_permission_required('manage_colors')
def delete_color(color_id):
    success, message = db.delete_color(color_id)
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    return redirect(url_for('manage_colors'))

# صفحات إدارة أنواع المنتجات
@app.route('/manage_product_types')
@page_permission_required('manage_product_types')
def manage_product_types():
    product_types = db.get_all_product_types()
    return render_template('manage_product_types.html', product_types=product_types)

@app.route('/add_product_type', methods=['POST'])
@action_permission_required('manage_product_types')
def add_product_type():
    type_name = request.form['type_name'].strip()
    if type_name:
        if db.add_product_type(type_name):
            flash(f'Product Type "{type_name}" added successfully!', 'success')
        else:
            flash(f'Error: Product Type "{type_name}" already exists!', 'error')
    return redirect(url_for('manage_product_types'))

@app.route('/edit_product_type/<int:type_id>', methods=['POST'])
@action_permission_required('manage_product_types')
def edit_product_type(type_id):
    new_name = request.form['type_name'].strip()
    if new_name:
        if db.update_product_type(type_id, new_name):
            flash(f'Product Type updated to "{new_name}" successfully!', 'success')
        else:
            flash('Error updating product type - name might already exist!', 'error')
    return redirect(url_for('manage_product_types'))

@app.route('/delete_product_type/<int:type_id>', methods=['POST'])
@action_permission_required('manage_product_types')
def delete_product_type(type_id):
    success, message = db.delete_product_type(type_id)
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    return redirect(url_for('manage_product_types'))

# صفحات إدارة فئات التجار
@app.route('/manage_trader_categories')
@page_permission_required('manage_trader_categories')
def manage_trader_categories():
    categories = db.get_all_trader_categories()
    return render_template('manage_trader_categories.html', categories=categories)

@app.route('/add_trader_category', methods=['POST'])
@action_permission_required('manage_trader_categories')
def add_trader_category():
    category_code = request.form['category_code'].strip().upper()
    category_name = request.form['category_name'].strip()
    description = request.form['description'].strip()
    
    if category_code and category_name:
        if db.add_trader_category(category_code, category_name, description):
            flash(f'Trader Category "{category_code}" added successfully!', 'success')
        else:
            flash(f'Error: Category code "{category_code}" already exists!', 'error')
    return redirect(url_for('manage_trader_categories'))

@app.route('/edit_trader_category/<int:category_id>', methods=['POST'])
@action_permission_required('manage_trader_categories')
def edit_trader_category(category_id):
    new_code = request.form['category_code'].strip().upper()
    new_name = request.form['category_name'].strip()
    new_description = request.form['description'].strip()
    
    if new_code and new_name:
        if db.update_trader_category(category_id, new_code, new_name, new_description):
            flash(f'Trader Category updated successfully!', 'success')
        else:
            flash('Error updating trader category!', 'error')
    return redirect(url_for('manage_trader_categories'))

@app.route('/delete_trader_category/<int:category_id>', methods=['POST'])
@action_permission_required('manage_trader_categories')
def delete_trader_category(category_id):
    success, message = db.delete_trader_category(category_id)
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    return redirect(url_for('manage_trader_categories'))

# صفحات إدارة Tags
@app.route('/manage_tags')
@page_permission_required('manage_tags')
def manage_tags():
    """صفحة إدارة Tags"""
    tags = db.get_all_tags()
    categories = db.get_tags_by_category()
    return render_template('manage_tags.html', tags=tags, categories=categories)

@app.route('/add_tag', methods=['POST'])
@action_permission_required('manage_tags')
def add_tag():
    """إضافة Tag جديد"""
    tag_name = request.form['tag_name'].strip()
    tag_category = request.form['tag_category'].strip()
    tag_color = request.form['tag_color']
    description = request.form['description'].strip()
    
    if tag_name:
        if db.add_tag(tag_name, tag_category, tag_color, description):
            flash(f'Tag "{tag_name}" added successfully!', 'success')
        else:
            flash(f'Error: Tag "{tag_name}" already exists!', 'error')
    return redirect(url_for('manage_tags'))

@app.route('/edit_tag/<int:tag_id>', methods=['POST'])
@action_permission_required('manage_tags')
def edit_tag(tag_id):
    """تعديل Tag"""
    new_name = request.form['tag_name'].strip()
    new_category = request.form['tag_category'].strip()
    new_color = request.form['tag_color']
    new_description = request.form['description'].strip()
    
    if new_name:
        if db.update_tag(tag_id, new_name, new_category, new_color, new_description):
            flash(f'Tag updated to "{new_name}" successfully!', 'success')
        else:
            flash('Error updating tag - name might already exist!', 'error')
    return redirect(url_for('manage_tags'))

@app.route('/delete_tag/<int:tag_id>', methods=['POST'])
@action_permission_required('manage_tags')
def delete_tag(tag_id):
    """حذف Tag"""
    success, message = db.delete_tag(tag_id)
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')
    return redirect(url_for('manage_tags'))

# صفحات المنتجات مع النظام المحدث
@app.route('/add_product_new', methods=['GET', 'POST'])
@page_permission_required('add_product')
def add_product_new():
    """صفحة إضافة منتج واحد مع المقاس والـ Tags والصور المنظمة"""
    if request.method == 'POST':
        try:
            product_code = request.form['product_code'].strip()
            brand_id = int(request.form['brand_id'])
            product_type_id = int(request.form['product_type_id'])
            trader_category = request.form['trader_category']
            product_size = request.form['product_size'].strip()
            wholesale_price = float(request.form['wholesale_price'])
            retail_price = float(request.form['retail_price'])
            initial_stock = int(request.form.get('initial_stock', 0))
            
            color_ids = request.form.getlist('color_ids')
            color_ids = [int(c) for c in color_ids if c]
            
            tag_ids = request.form.getlist('tag_ids')
            tag_ids = [int(t) for t in tag_ids if t]
            
            if not color_ids:
                flash('Please select at least one color!', 'error')
                return redirect(url_for('add_product_new'))
            
            if db.check_product_exists(product_code, brand_id, trader_category):
                flash('Product with same code, brand, and category already exists!', 'error')
                return redirect(url_for('add_product_new'))
            
            # إضافة المنتج مع المقاس والـ Tags
            success, result = db.add_base_product_with_variants(
                product_code, brand_id, product_type_id, trader_category, product_size,
                wholesale_price, retail_price, color_ids, tag_ids, initial_stock
            )
            
            if success:
                base_product_id = result
                
                # حفظ لينكات الصور لكل لون
                uploaded_images = 0
                for color_id in color_ids:
                    image_url_key = f'color_image_url_{color_id}'
                    image_url = request.form.get(image_url_key, '').strip()
                    
                    if image_url:
                        # جلب variant_id
                        conn = db.get_connection()
                        cursor = conn.cursor()
                        cursor.execute('''
                            SELECT id FROM product_variants
                            WHERE base_product_id = ? AND color_id = ?
                        ''', (base_product_id, color_id))
                        variant_result = cursor.fetchone()
                        conn.close()
                        
                        if variant_result:
                            variant_id = variant_result[0]
                            if db.add_color_image(variant_id, image_url):
                                uploaded_images += 1
                
                flash(f'Product "{product_code}" added successfully with {len(color_ids)} colors, {len(tag_ids)} tags and {uploaded_images} image URLs!', 'success')
                return redirect(url_for('products_new'))
            else:
                flash(f'Error adding product: {result}', 'error')
                
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
    
    brands = db.get_all_brands()
    colors = db.get_all_colors()
    product_types = db.get_all_product_types()
    trader_categories = db.get_all_trader_categories()
    tags = db.get_all_tags()
    
    return render_template('add_product_new.html', 
                         brands=brands, 
                         colors=colors, 
                         product_types=product_types,
                         trader_categories=trader_categories,
                         tags=tags)

@app.route('/add_products_multi', methods=['GET', 'POST'])
@page_permission_required('add_multiple')
def add_products_multi():
    """صفحة إضافة منتجات متعددة مع المقاس والـ Tags"""
    if request.method == 'POST':
        try:
            num_products = int(request.form.get('num_products', 1))
            
            products_data = []
            
            for i in range(num_products):
                product_code = request.form.get(f'product_code_{i}', '').strip()
                if not product_code:
                    continue
                
                brand_id = request.form.get(f'brand_id_{i}')
                product_type_id = request.form.get(f'product_type_id_{i}')
                trader_category = request.form.get(f'trader_category_{i}')
                product_size = request.form.get(f'product_size_{i}', '').strip()
                wholesale_price = request.form.get(f'wholesale_price_{i}')
                retail_price = request.form.get(f'retail_price_{i}')
                initial_stock = request.form.get(f'initial_stock_{i}', 0)
                
                color_ids = request.form.getlist(f'color_ids_{i}')
                color_ids = [int(c) for c in color_ids if c]
                
                tag_ids = request.form.getlist(f'tag_ids_{i}')
                tag_ids = [int(t) for t in tag_ids if t]
                
                if not color_ids:
                    flash(f'Product {i+1}: Please select at least one color!', 'error')
                    continue
                
                if not brand_id or not product_type_id or not trader_category or not wholesale_price or not retail_price:
                    flash(f'Product {i+1}: Please fill all required fields!', 'error')
                    continue
                
                products_data.append({
                    'product_code': product_code,
                    'brand_id': int(brand_id),
                    'product_type_id': int(product_type_id),
                    'trader_category': trader_category,
                    'product_size': product_size,
                    'wholesale_price': float(wholesale_price),
                    'retail_price': float(retail_price),
                    'initial_stock': int(initial_stock),
                    'color_ids': color_ids,
                    'tag_ids': tag_ids
                })
            
            if not products_data:
                flash('No valid products to add!', 'error')
                return redirect(url_for('add_products_multi'))
            
            result = db.add_multiple_products_batch(products_data)
            
            if result['success']:
                if result['success_count'] > 0:
                    flash(f'Successfully added {result["success_count"]} products!', 'success')
                
                if result['failed_count'] > 0:
                    flash(f'{result["failed_count"]} products failed to add. Check details below.', 'warning')
                    for failed in result['failed_products']:
                        flash(f'Failed to add product {failed["product"]["product_code"]}: {failed["error"]}', 'error')
                
                return redirect(url_for('products_new'))
            else:
                flash(f'Error adding products: {result["error"]}', 'error')
                
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
    
    brands = db.get_all_brands()
    colors = db.get_all_colors()
    product_types = db.get_all_product_types()
    trader_categories = db.get_all_trader_categories()
    tags = db.get_all_tags()
    
    return render_template('add_products_multi.html', 
                         brands=brands, 
                         colors=colors, 
                         product_types=product_types,
                         trader_categories=trader_categories,
                         tags=tags)

@app.route('/products_new')
@page_permission_required('products')
def products_new():
    """صفحة عرض المنتجات المحسنة مع المقاس والـ Tags والصور"""
    search_term = request.args.get('search', '')
    products = db.get_products_with_color_images(search_term)
    return render_template('products_new.html', products=products, search_term=search_term)

@app.route('/search_products')
@login_required
def search_products():
    """البحث في المنتجات - AJAX مع المقاس والـ Tags"""
    search_term = request.args.get('q', '')
    products = db.get_products_with_color_images(search_term)
    
    results = []
    for product in products:
        # تحضير بيانات الألوان مع المخزون
        colors_with_stock = []
        for color in product[10]:  # colors_data
            colors_with_stock.append(f"{color['name']}: {color['stock']}")
        
        # تحضير بيانات Tags
        tags_list = [tag[1] for tag in product[12]] if product[12] else []
        
        results.append({
            'id': product[0],
            'code': product[1],
            'brand': product[2],
            'type': product[3],
            'category': product[4],
            'size': product[5],
            'wholesale': product[6],
            'retail': product[7],
            'supplier': product[8],
            'colors_data': product[10],
            'colors_text': ', '.join(colors_with_stock),
            'total_stock': product[11],
            'tags': tags_list,
            'created': product[9][:10] if product[9] else 'N/A'
        })
    
    return jsonify({'products': results})

@app.route('/product_details/<int:product_id>')
@page_permission_required('product_details')
def product_details(product_id):
    """صفحة تفاصيل منتج واحد مع الصور والمقاس والـ Tags"""
    details = db.get_product_details(product_id)
    
    if not details:
        flash('Product not found!', 'error')
        return redirect(url_for('products_new'))
    
    return render_template('product_details.html', details=details)

@app.route('/delete_product/<int:product_id>', methods=['POST'])
@action_permission_required('products')
def delete_product(product_id):
    """حذف منتج مع كل بياناته"""
    success, message = db.delete_product(product_id)
    
    if success:
        flash(message, 'success')
    else:
        flash(f'Error: {message}', 'error')
    
    return redirect(url_for('products_new'))

# صفحات الجرد الشامل مع المقاس والـ Tags
@app.route('/inventory_management')
@page_permission_required('bulk_inventory')
def inventory_management():
    """صفحة الجرد الشاملة مع المقاس والـ Tags"""
    search_term = request.args.get('search', '')
    brand_filter = request.args.get('brand', '')
    category_filter = request.args.get('category', '')
    
    inventory_data = db.get_all_products_for_inventory(search_term, brand_filter, category_filter)
    summary = db.get_inventory_summary()
    brands = db.get_brands_for_filter()
    categories = db.get_categories_for_filter()
    
    return render_template('inventory_management.html', 
                         inventory_data=inventory_data,
                         summary=summary,
                         brands=brands,
                         categories=categories,
                         search_term=search_term,
                         brand_filter=brand_filter,
                         category_filter=category_filter)

@app.route('/update_inventory', methods=['POST'])
@action_permission_required('bulk_inventory')
def update_inventory():
    """تحديث المخزون بكميات جديدة"""
    try:
        # تجميع كل التحديثات
        stock_updates = []
        
        for key, value in request.form.items():
            if key.startswith('stock_'):
                variant_id = int(key.replace('stock_', ''))
                new_stock = int(value)
                stock_updates.append({
                    'variant_id': variant_id,
                    'new_stock': new_stock
                })
        
        # Get details for each variant before update (for logging)
        conn = db.get_connection()
        cursor = conn.cursor()
        
        logged_count = 0
        for update in stock_updates:
            try:
                cursor.execute('''
                    SELECT pv.current_stock, bp.id, bp.product_code, b.brand_name,
                           pt.type_name, c.color_name, ci.image_url
                    FROM product_variants pv
                    JOIN base_products bp ON pv.base_product_id = bp.id
                    JOIN brands b ON bp.brand_id = b.id
                    JOIN product_types pt ON bp.product_type_id = pt.id
                    JOIN colors c ON pv.color_id = c.id
                    LEFT JOIN color_images ci ON pv.id = ci.variant_id
                    WHERE pv.id = ?
                ''', (update['variant_id'],))
                result = cursor.fetchone()
                
                if result and result[0] != update['new_stock']:  # Only log if changed
                    update['old_stock'] = result[0]
                    update['product_id'] = result[1]
                    update['product_code'] = result[2]
                    update['brand_name'] = result[3]
                    update['product_type'] = result[4]
                    update['color_name'] = result[5]
                    update['image_url'] = result[6] or ''
                else:
                    update['skip_log'] = True
            except:
                update['skip_log'] = True
        
        conn.close()
        
        # استخدام الدالة الموجودة أصلاً
        result = db.bulk_update_inventory(stock_updates)
        
        # Log successful updates
        for update in stock_updates:
            if not update.get('skip_log', False) and 'old_stock' in update:
                db.add_stock_log(
                    operation_type='Bulk Update',
                    product_id=update['product_id'],
                    variant_id=update['variant_id'],
                    product_code=update['product_code'],
                    brand_name=update['brand_name'],
                    product_type=update['product_type'],
                    color_name=update['color_name'],
                    image_url=update['image_url'],
                    old_value=update['old_stock'],
                    new_value=update['new_stock'],
                    username='Admin',
                    notes='Bulk inventory update',
                    source_page='Inventory Management',
                    source_url=request.url
                )
                logged_count += 1
        
        # تحقق من AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if is_ajax:
            if result['success']:
                return jsonify({
                    'success': True,
                    'updated_count': result['updated_count'],
                    'logged_count': logged_count,
                    'message': f'Successfully updated {result["updated_count"]} product(s)'
                }), 200
            else:
                return jsonify({
                    'success': False,
                    'error': result.get('error', 'Unknown error')
                }), 500
        else:
            if result['success']:
                flash(f'Successfully updated {result["updated_count"]} product(s)!', 'success')
            else:
                flash(f'Error updating inventory: {result.get("error", "Unknown error")}', 'error')
            
            return redirect(url_for('inventory_management', 
                                  search=request.args.get('search', ''),
                                  brand=request.args.get('brand', ''),
                                  category=request.args.get('category', '')))
        
    except Exception as e:
        print(f"Error in update_inventory: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': str(e)}), 500
        else:
            flash(f'Error updating inventory: {str(e)}', 'error')
            return redirect(url_for('inventory_management'))

@app.route('/inventory_search')
@login_required
def inventory_search():
    """البحث في صفحة الجرد - AJAX"""
    search_term = request.args.get('q', '')
    brand_filter = request.args.get('brand', '')
    category_filter = request.args.get('category', '')
    
    inventory_data = db.get_all_products_for_inventory(search_term, brand_filter, category_filter)
    
    results = []
    for item in inventory_data:
        product = item['product']
        color_variants = item['color_variants']
        
        results.append({
            'product_id': product[0],
            'product_code': product[1],
            'brand': product[2],
            'type': product[3],
            'category': product[4],
            'size': product[5],
            'total_stock': item['total_stock'],
            'tags': [tag[1] for tag in item['tags']],
            'color_variants': [
                {
                    'variant_id': cv[0],
                    'color_id': cv[1],
                    'color_name': cv[2],
                    'color_code': cv[3],
                    'current_stock': cv[4]
                } for cv in color_variants
            ]
        })
    
    return jsonify({'inventory_data': results})

# صفحات Excel Bulk Upload مع النظام المحدث
@app.route('/bulk_upload_excel', methods=['GET', 'POST'])
@page_permission_required('bulk_upload')
def bulk_upload_excel():
    """Excel Bulk Upload"""
    if request.method == 'POST':
        try:
            # Check if file exists
            if 'excel_file' not in request.files:
                flash('لم يتم اختيار ملف!', 'error')
                return redirect(url_for('bulk_upload_excel'))
            
            file = request.files['excel_file']
            
            if not file or not file.filename:
                flash('يرجى اختيار ملف Excel!', 'error')
                return redirect(url_for('bulk_upload_excel'))
            
            # Check file extension
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                flash('يرجى رفع ملف Excel بصيغة .xlsx أو .xls!', 'error')
                return redirect(url_for('bulk_upload_excel'))
            
            print(f"📁 Processing file: {file.filename}")
            
            # ✅ Read Excel with openpyxl (instead of pandas)
            
            wb = load_workbook(file, data_only=True)
            sheet = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1] if cell.value]
            
            print(f"📊 Found {len(headers)} columns")
            
            # Check required columns
            required_columns = [
                'Product Code', 'Brand Name', 'Product Type', 'Category',
                'Wholesale Price', 'Retail Price', 'Color Name', 'Stock'
            ]
            
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                flash(f'الأعمدة المطلوبة مفقودة: {", ".join(missing_columns)}', 'error')
                return redirect(url_for('bulk_upload_excel'))
            
            # Convert rows to list of dictionaries
            excel_data = []
            row_count = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not any(row):
                    continue
                
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        value = row[idx]
                        # Convert to string and handle None
                        row_dict[header] = str(value).strip() if value is not None else ''
                    else:
                        row_dict[header] = ''
                
                excel_data.append(row_dict)
                row_count += 1
            
            print(f"📝 Loaded {row_count} rows")
            
            # Limit rows to prevent timeout
            MAX_ROWS = 2000
            if len(excel_data) > MAX_ROWS:
                flash(f'⚠️ تحذير! الملف يحتوي على {len(excel_data)} صف. سيتم تحميل أول {MAX_ROWS} صف فقط.', 'warning')
                excel_data = excel_data[:MAX_ROWS]
                flash(f'تم تقليل البيانات إلى {MAX_ROWS} صف.', 'info')
            
            # Process the data
            result = db.bulk_add_products_from_excel_enhanced(excel_data)
            
            if result['success']:
                # Backup after successful upload
                print("📦 Creating backup after bulk upload...")
                time.sleep(2)
                backup_success = backup_system.create_backup()
                if backup_success:
                    print("✅ Backup created successfully!")
                else:
                    print("❌ Backup failed!")
                
                # Success message
                success_msg = f"✅ تم رفع {result['success_count']} منتج من أصل {len(excel_data)} بنجاح!"
                flash(success_msg, 'success')
                
                # Show created items
                if result['created_brands']:
                    flash(f"تم إضافة الماركات: {', '.join(result['created_brands'])}", 'info')
                
                if result['created_colors']:
                    flash(f"تم إضافة الألوان: {', '.join(result['created_colors'])}", 'info')
                
                if result['created_types']:
                    flash(f"تم إضافة الأنواع: {', '.join(result['created_types'])}", 'info')
                
                # Show errors if any
                if result['failed_count'] > 0:
                    flash(f'⚠️ فشل رفع {result["failed_count"]} منتج. يرجى مراجعة الأخطاء أدناه.', 'warning')
                
                # Show first 5 errors
                for failed in result['failed_products'][:5]:
                    flash(f"❌ صف {failed['row']}: {failed['error']}", 'error')
                
                if len(result['failed_products']) > 5:
                    flash(f"⚠️ و {len(result['failed_products']) - 5} أخطاء أخرى...", 'warning')
                
                return redirect(url_for('products_new'))
            
            else:
                error_msg = result.get('error', 'حدث خطأ غير متوقع')
                flash(f'❌ {error_msg}', 'error')
                print(f"❌ Bulk upload error: {error_msg}")
        
        except InvalidFileException:
            flash('❌ الملف تالف أو ليس ملف Excel صالح!', 'error')
        
        except Exception as e:
            error_msg = str(e)
            flash(f'❌ خطأ: {error_msg}', 'error')
            print(f"❌ Exception in bulk upload: {error_msg}")
    
    return render_template('bulk_upload_excel.html')

@app.route('/export_products', methods=['GET', 'POST'])
@page_permission_required('export_products')
def export_products():
    """صفحة تصدير المنتجات مع فلاتر متعددة"""
    if request.method == 'POST':
        try:
            # جلب الفلاتر من الـ form
            selected_brands = request.form.getlist('brands')
            selected_categories = request.form.getlist('categories')
            selected_types = request.form.getlist('product_types')
            selected_colors = request.form.getlist('colors')
            selected_products = request.form.getlist('product_codes')
            stock_filter = request.form.get('stock_filter', 'all')
            
            # جلب كل المنتجات
            all_products = db.get_products_with_color_images('')
            
            # تطبيق الفلاتر
            filtered_data = []
            
            for product in all_products:
                product_id = product[0]
                product_code = product[1]
                brand_name = product[2]
                product_type = product[3]
                category = product[4]
                size = product[5]
                wholesale = product[6]
                retail = product[7]
                colors_data = product[10]  # قائمة الألوان مع التفاصيل
                tags = product[12] if len(product) > 12 else []
                
                # تطبيق فلاتر Brand, Category, Type, Product Code
                if selected_brands and brand_name not in selected_brands:
                    continue
                if selected_categories and category not in selected_categories:
                    continue
                if selected_types and product_type not in selected_types:
                    continue
                if selected_products and product_code not in selected_products:
                    continue
                
                # معالجة كل لون كصف منفصل
                for color_info in colors_data:
                    color_name = color_info['name']
                    color_stock = color_info['stock']
                    image_url = color_info['image_url'] or ''
                    
                    # تطبيق فلتر اللون
                    if selected_colors and color_name not in selected_colors:
                        continue
                    
                    # تطبيق فلتر المخزون
                    if stock_filter == 'in_stock' and color_stock <= 0:
                        continue
                    elif stock_filter == 'out_of_stock' and color_stock > 0:
                        continue
                    elif stock_filter == 'low_stock' and (color_stock > 5 or color_stock <= 0):
                        continue
                    
                    # تجهيز Tags كنص (مفصول بفواصل)
                    tags_text = ','.join([tag[1] for tag in tags]) if tags else ''
                    
                    # إضافة الصف
                    filtered_data.append({
                        'Product Code': product_code,
                        'Brand Name': brand_name,
                        'Product Type': product_type,
                        'Category': category,
                        'Size': size or '',
                        'Wholesale Price': wholesale,
                        'Retail Price': retail,
                        'Color Name': color_name,
                        'Stock': color_stock,
                        'Image URL': image_url,
                        'Tags': tags_text
                    })
            
            if not filtered_data:
                flash('No products match the selected filters!', 'warning')
                return redirect(url_for('export_products'))
            
            # ✅ إنشاء Excel باستخدام openpyxl
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # ✅ كتابة الـ Headers
            headers = ['Product Code', 'Brand Name', 'Product Type', 'Category', 'Size', 
                      'Wholesale Price', 'Retail Price', 'Color Name', 'Stock', 'Image URL', 'Tags']
            ws.append(headers)
            
            # ✅ تنسيق الـ Headers (bold)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            # ✅ كتابة البيانات
            for row_data in filtered_data:
                ws.append([
                    row_data['Product Code'],
                    row_data['Brand Name'],
                    row_data['Product Type'],
                    row_data['Category'],
                    row_data['Size'],
                    row_data['Wholesale Price'],
                    row_data['Retail Price'],
                    row_data['Color Name'],
                    row_data['Stock'],
                    row_data['Image URL'],
                    row_data['Tags']
                ])
            
            # ✅ حفظ في BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # تحديد اسم الملف حسب الفلاتر
            filename = f'products_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            flash(f'Exported {len(filtered_data)} product variants successfully!', 'success')
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            flash(f'Error exporting products: {str(e)}', 'error')
            return redirect(url_for('export_products'))
    
    # GET request - عرض الصفحة مع الفلاتر
    brands = db.get_brands_for_filter()
    categories = db.get_categories_for_filter()
    product_types = [pt[1] for pt in db.get_all_product_types()]
    colors = [c[1] for c in db.get_all_colors()]
    
    # جلب كل أكواد المنتجات
    all_products = db.get_products_with_color_images('')
    product_codes = sorted(list(set([p[1] for p in all_products])))
    
    return render_template('export_products.html',
                         brands=brands,
                         categories=categories,
                         product_types=product_types,
                         colors=colors,
                         product_codes=product_codes)


@app.route('/download_excel_template')
@login_required
def download_excel_template():
    """تحميل نموذج Excel المحسن - نسخة مبسطة بدون تنسيق"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        # ✅ البيانات التجريبية
        template_data = [
            ['Product Code', 'Brand Name', 'Product Type', 'Category', 'Size', 'Wholesale Price', 'Retail Price', 'Color Name', 'Stock', 'Image URL', 'Tags'],
            ['96115', 'Tommy Hilfiger', 'Handbag', 'L', '20×22×5', 1000, 1500, 'Black', 15, 'https://images.unsplash.com/photo-1553062407-98eeb64c6a62?w=400', 'Sale,Medium'],
            ['96115', 'Tommy Hilfiger', 'Handbag', 'L', '20×22×5', 1000, 1500, 'Red', 5, 'https://images.unsplash.com/photo-1584917865442-de89df76afd3?w=400', 'Sale,Medium'],
            ['96115', 'Tommy Hilfiger', 'Handbag', 'L', '20×22×5', 1000, 1500, 'Brown', 3, 'https://images.unsplash.com/photo-1594633312681-425c7b97ccd1?w=400', 'Sale,Medium'],
            ['87432', 'Gucci', 'Wallet', 'F', '15×18×3', 1200, 1800, 'Gold', 10, 'https://images.unsplash.com/photo-1548036328-c9fa89d128fa?w=400', 'New Arrival,Small'],
            ['87432', 'Gucci', 'Wallet', 'F', '15×18×3', 1200, 1800, 'Silver', 8, 'https://images.unsplash.com/photo-1553062407-98eeb64c6a62?w=400', 'New Arrival,Small'],
            ['75321', 'Zara', 'Backpack', 'L', '25×30×10', 800, 1200, 'Navy Blue', 20, 'https://images.unsplash.com/photo-1622560480605-d83c853bc5c3?w=400', 'Summer,Large']
        ]
        
        # ✅ إنشاء Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Products Template"
        
        # ✅ كتابة كل الصفوف (Headers + Data)
        for row in template_data:
            ws.append(row)
        
        # ✅ تنسيق الـ Headers (أول صف - bold)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # ✅ حفظ في BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='products_template_enhanced_v2.xlsx'
        )
        
    except Exception as e:
        print(f"Error creating template: {e}")
        flash('Error creating template file', 'error')
        return redirect(url_for('bulk_upload_excel'))

@app.route('/edit_product/<int:product_id>', methods=['GET', 'POST'])
@page_permission_required('edit_product')
def edit_product(product_id):
    """تعديل بيانات المنتج الأساسية - كل البيانات قابلة للتعديل"""
    if request.method == 'POST':
        try:
            # جلب البيانات الجديدة (كلها قابلة للتعديل)
            product_code = request.form['product_code'].strip()
            brand_id = int(request.form['brand_id'])
            product_type_id = int(request.form['product_type_id'])
            trader_category = request.form['trader_category']
            product_size = request.form.get('product_size', '').strip()
            wholesale_price = float(request.form['wholesale_price'])
            retail_price = float(request.form['retail_price'])
            
            # التحقق من عدم التكرار (إذا تغير الكود/البراند/الفئة)
            old_details = db.get_product_details(product_id)
            old_code = old_details['product'][1]
            old_category = old_details['product'][2]
            
            # جلب brand_id القديم
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT brand_id FROM base_products WHERE id = ?', (product_id,))
            old_brand_id = cursor.fetchone()[0]
            conn.close()
            
            # إذا تغير الكود أو البراند أو الفئة، نتحقق من التكرار
            if (product_code != old_code or brand_id != old_brand_id or trader_category != old_category):
                if db.check_product_exists(product_code, brand_id, trader_category):
                    flash('Product with same code, brand, and category already exists!', 'error')
                    return redirect(url_for('edit_product', product_id=product_id))
            
            # تحديث المنتج
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE base_products 
                SET product_code = ?, brand_id = ?, product_type_id = ?, 
                    trader_category = ?, product_size = ?, wholesale_price = ?, retail_price = ?
                WHERE id = ?
            ''', (product_code, brand_id, product_type_id, trader_category, 
                  product_size, wholesale_price, retail_price, product_id))
            conn.commit()
            conn.close()
            
            flash('Product updated successfully!', 'success')
            return redirect(url_for('product_details', product_id=product_id))
            
        except Exception as e:
            flash(f'Error updating product: {str(e)}', 'error')
    
    # جلب بيانات المنتج للعرض في النموذج
    details = db.get_product_details(product_id)
    if not details:
        flash('Product not found!', 'error')
        return redirect(url_for('products_new'))
    
    brands = db.get_all_brands()
    product_types = db.get_all_product_types()
    trader_categories = db.get_all_trader_categories()
    tags = db.get_all_tags()
    
    return render_template('edit_product.html', 
                         details=details, 
                         brands=brands,
                         product_types=product_types,
                         trader_categories=trader_categories,
                         tags=tags)


@app.route('/update_stock/<int:variant_id>', methods=['POST'])
@action_permission_required('product_details')
def update_stock(variant_id):
    """تحديث مخزون لون معين"""
    try:
        new_stock = int(request.form['new_stock'])
        product_id = int(request.form.get('product_id'))
        
        # Get old stock + product details for logging
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT pv.current_stock, bp.product_code, b.brand_name, 
                   pt.type_name, c.color_name, ci.image_url, bp.id
            FROM product_variants pv
            JOIN base_products bp ON pv.base_product_id = bp.id
            JOIN brands b ON bp.brand_id = b.id
            JOIN product_types pt ON bp.product_type_id = pt.id
            JOIN colors c ON pv.color_id = c.id
            LEFT JOIN color_images ci ON pv.id = ci.variant_id
            WHERE pv.id = ?
        ''', (variant_id,))
        result = cursor.fetchone()
        
        if result:
            old_stock = result[0]
            product_code = result[1]
            brand_name = result[2]
            product_type = result[3]
            color_name = result[4]
            image_url = result[5] or ''
            product_id = result[6]
            
            # Update stock
            cursor.execute('UPDATE product_variants SET current_stock = ? WHERE id = ?', 
                          (new_stock, variant_id))
            conn.commit()
            conn.close()
            
            # Log the operation
            db.add_stock_log(
                operation_type='Stock Update',
                product_id=product_id,
                variant_id=variant_id,
                product_code=product_code,
                brand_name=brand_name,
                product_type=product_type,
                color_name=color_name,
                image_url=image_url,
                old_value=old_stock,
                new_value=new_stock,
                username='Admin',
                notes=f'Manual update from product details',
                source_page='Product Details',
                source_url=request.url
            )
            
            flash('Stock updated successfully!', 'success')
        else:
            conn.close()
            flash('Variant not found!', 'error')
        
    except Exception as e:
        flash(f'Error updating stock: {str(e)}', 'error')
    
    return redirect(url_for('product_details', product_id=product_id))

@app.route('/upload_color_image/<int:variant_id>', methods=['POST'])
@action_permission_required('product_details')
def upload_color_image(variant_id):
    """تحديث لينك صورة اللون"""
    try:
        image_url = request.form.get('image_url', '').strip()
        
        if not image_url:
            flash('Please enter an image URL!', 'error')
            return redirect(request.referrer)
        
        # Get product details for logging
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT bp.id, bp.product_code, b.brand_name, pt.type_name, c.color_name, ci.image_url
            FROM product_variants pv
            JOIN base_products bp ON pv.base_product_id = bp.id
            JOIN brands b ON bp.brand_id = b.id
            JOIN product_types pt ON bp.product_type_id = pt.id
            JOIN colors c ON pv.color_id = c.id
            LEFT JOIN color_images ci ON pv.id = ci.variant_id
            WHERE pv.id = ?
        ''', (variant_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            product_id = result[0]
            product_code = result[1]
            brand_name = result[2]
            product_type = result[3]
            color_name = result[4]
            old_image_url = result[5]
            
            # Update image
            if db.add_color_image(variant_id, image_url):
                # Log the operation
                db.add_stock_log(
                    operation_type='Image Updated',
                    product_id=product_id,
                    variant_id=variant_id,
                    product_code=product_code,
                    brand_name=brand_name,
                    product_type=product_type,
                    color_name=color_name,
                    image_url=image_url,
                    old_value=None,
                    new_value=None,
                    username='Admin',
                    notes=f'Image {"updated" if old_image_url else "added"}',
                    source_page='Product Details',
                    source_url=request.referrer or ''
                )
                
                flash('Image URL updated successfully!', 'success')
            else:
                flash('Error updating image URL!', 'error')
        else:
            flash('Variant not found!', 'error')
            
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(request.referrer)

@app.route('/logs')
@page_permission_required('activity_logs')
def logs():
    """Stock Activity Logs"""
    # Get filters
    operation_filter = request.args.get('operation', '')
    date_from = request.args.get('datefrom', '')
    date_to = request.args.get('dateto', '')
    search_term = request.args.get('search', '')
    limit = int(request.args.get('limit', 100))
    
    # Get logs
    logs = db.get_all_logs(
        limit=limit,
        operation_filter=operation_filter if operation_filter and operation_filter != 'All' else None,
        date_from=date_from if date_from else None,
        date_to=date_to if date_to else None,
        search_term=search_term if search_term else None
    )
    
    # Get statistics
    stats = db.get_logs_stats()
    
    # ✅ أضف كل الـ operation types المتاحة
    operation_types = [
        'All',
        'Stock Update',
        'Bulk Update', 
        'Product Added',
        'Product Deleted',
        'Image Updated',
        'Barcode Generated',           # ✅ جديد
        'Barcode Labels Printed',      # ✅ جديد
        'Barcode Scan - Add',          # ✅ جديد
        'Barcode Scan - Remove'        # ✅ جديد
    ]
    
    return render_template('logs.html', 
                          logs=logs, 
                          stats=stats, 
                          operation_types=operation_types,
                          operation_filter=operation_filter,
                          date_from=date_from,
                          date_to=date_to,
                          search_term=search_term)

@app.route('/export_logs')
@action_permission_required('activity_logs')
def export_logs():
    """Export logs to Excel"""
    try:
                
        # Get all logs
        logs = db.get_all_logs(limit=10000)  # Get more for export
        
        # ✅ إنشاء Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Activity Logs"
        
        # ✅ كتابة الـ Headers
        headers = ['Date', 'Time', 'Product Code', 'Brand', 'Type', 'Color', 
                  'Old Stock', 'New Stock', 'Change', 'Operation', 'Source Page', 
                  'User', 'Image URL']
        ws.append(headers)
        
        # ✅ تنسيق الـ Headers (bold)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # ✅ كتابة البيانات
        for log in logs:
            # Parse created_date
            created_date = log[16] if log[16] else ''
            if created_date:
                try:
                    dt = datetime.fromisoformat(str(created_date))
                    date_str = dt.strftime('%Y-%m-%d')
                    time_str = dt.strftime('%H:%M:%S')
                except:
                    date_str = str(created_date)[:10]
                    time_str = str(created_date)[11:19]
            else:
                date_str = 'N/A'
                time_str = 'N/A'
            
            # إضافة الصف
            ws.append([
                date_str,
                time_str,
                log[4] or 'N/A',                           # Product Code
                log[5] or 'N/A',                           # Brand
                log[6] or 'N/A',                           # Type
                log[7] or 'N/A',                           # Color
                log[9] if log[9] is not None else '-',    # Old Stock
                log[10] if log[10] is not None else '-',  # New Stock
                log[11] if log[11] is not None else '-',  # Change
                log[1] or 'N/A',                           # Operation
                log[14] or 'N/A',                          # Source Page
                log[12] or 'Admin',                        # User
                log[8] or ''                               # Image URL
            ])
        
        # ✅ حفظ في BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'stock_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting logs: {str(e)}', 'error')
        return redirect(url_for('logs'))


# ====================================================================
# BARCODE SYSTEM ROUTES
# ====================================================================

# === Barcode Management Page ===

@app.route('/barcode/management')
@page_permission_required('barcode_system')
def barcode_management():
    """Barcode management page"""
    
    # Get filters
    search = request.args.get('search', '')
    brand_filter = request.args.get('brand', '')
    type_filter = request.args.get('type', '')
    color_filter = request.args.get('color', '')
    tab = request.args.get('tab', 'without')
    image_filter = request.args.get('image_filter', 'all')  # للـ With Barcode
    stock_filter = request.args.get('stock_filter', '')  # ← NEW: للـ Without Barcode
    page = int(request.args.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page
    
    # Get statistics
    stats = db.get_barcode_stats()
    
    # Get image status stats (للـ With Barcode)
    image_stats = db.count_barcode_image_status()
    stats.update(image_stats)
    
    # Get filter options
    brands = db.get_brands_for_filter()
    types = [pt[1] for pt in db.get_all_product_types()]
    colors = [c[1] for c in db.get_all_colors()]
    
    # Get variants based on tab
    if tab == 'without':
        # ✅ تطبيق الـ stock filter
        variants = db.get_variants_without_barcode(
            search=search,
            brand_filter=brand_filter,
            type_filter=type_filter,
            color_filter=color_filter,
            limit=per_page,
            offset=offset,
            in_stock_only=(stock_filter == 'in_stock')  # ← NEW
        )
        total_count = db.count_variants_without_barcode(
            search=search,
            brand_filter=brand_filter,
            type_filter=type_filter,
            color_filter=color_filter,
            in_stock_only=(stock_filter == 'in_stock')  # ← NEW
        )
    else:
        variants = db.get_barcodes_with_image_status(
            search=search,
            brand_filter=brand_filter,
            type_filter=type_filter,
            color_filter=color_filter,
            image_filter=image_filter,
            limit=per_page,
            offset=offset
        )
        total_count = len(variants)
    
    total_pages = (total_count + per_page - 1) // per_page
    has_more = page < total_pages
    
    return render_template('barcode_management.html',
        stats=stats,
        variants=variants,
        brands=brands,
        types=types,
        colors=colors,
        tab=tab,
        image_filter=image_filter,
        search=search,
        brand_filter=brand_filter,
        type_filter=type_filter,
        color_filter=color_filter,
        page=page,
        total_pages=total_pages,
        total_count=total_count,
        has_more=has_more
    )


@app.route('/barcode/generate/<int:variant_id>', methods=['POST'])
@action_permission_required('barcode_system')
def generate_barcode_single(variant_id):
    """Generate barcode for a single variant"""
    try:
        # Check if barcode already exists
        existing = db.get_barcode_by_variant(variant_id)
        if existing:
            return jsonify({
                'success': False,
                'error': 'Barcode already exists for this variant'
            }), 400
        
        # Get variant details
        variant = db.get_variant_details_for_barcode(variant_id)
        if not variant:
            return jsonify({
                'success': False,
                'error': 'Variant not found'
            }), 404
        
        product_code = variant[1]
        color_name = variant[4]
        
        # Generate barcode
        result = generate_barcode_for_variant(product_code, color_name, variant_id)
        
        if not result:
            return jsonify({
                'success': False,
                'error': 'Failed to generate barcode'
            }), 500
        
        # Save to database
        user_id = session.get('user_id', 0)
        barcode_id = db.create_barcode(
            variant_id=variant_id,
            barcode_number=result['barcode'],
            image_path=result['image_path'],
            user_id=user_id
        )
        
        if not barcode_id:
            return jsonify({
                'success': False,
                'error': 'Failed to save barcode to database'
            }), 500
        
        # Log the operation
        db.add_stock_log(
            operation_type='Barcode Generated',
            product_id=None,
            variant_id=variant_id,
            product_code=product_code,
            brand_name=variant[2],
            product_type=variant[3],
            color_name=color_name,
            image_url=variant[7] or '',
            old_value=None,
            new_value=None,
            username=session.get('full_name', 'User'),
            notes=f'Barcode {result["barcode"]} generated',
            source_page='Barcode Management',
            source_url=request.url
        )
        
        return jsonify({
            'success': True,
            'barcode': result['barcode'],
            'image_path': result['image_path'],
            'message': f'Barcode generated successfully for {product_code} - {color_name}'
        })
        
    except Exception as e:
        print(f"❌ Error generating barcode: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/barcode/generate_all', methods=['POST'])
@action_permission_required('barcode_system')
def generate_all_barcodes():
    """Generate barcodes for all variants without barcodes"""
    try:
        # Get filters from request
        data = request.get_json() or {}
        search = data.get('search', '')
        brand_filter = data.get('brand', '')
        type_filter = data.get('type', '')
        color_filter = data.get('color', '')
        
        # Get all variants without barcodes (no limit)
        variants = db.get_variants_without_barcode(
            search=search,
            brand_filter=brand_filter,
            type_filter=type_filter,
            color_filter=color_filter,
            limit=10000,  # Large limit
            offset=0
        )
        
        if not variants:
            return jsonify({
                'success': False,
                'error': 'No variants found to generate barcodes'
            }), 400
        
        success_count = 0
        failed_count = 0
        failed_variants = []
        user_id = session.get('user_id', 0)
        
        for variant in variants:
            variant_id = variant[0]
            product_code = variant[1]
            color_name = variant[4]
            
            try:
                # Generate barcode
                result = generate_barcode_for_variant(product_code, color_name, variant_id)
                
                if result:
                    # Save to database
                    barcode_id = db.create_barcode(
                        variant_id=variant_id,
                        barcode_number=result['barcode'],
                        image_path=result['image_path'],
                        user_id=user_id
                    )
                    
                    if barcode_id:
                        success_count += 1
                        
                        # Log the operation
                        db.add_stock_log(
                            operation_type='Barcode Generated',
                            product_id=None,
                            variant_id=variant_id,
                            product_code=product_code,
                            brand_name=variant[2],
                            product_type=variant[3],
                            color_name=color_name,
                            image_url=variant[7] or '',
                            old_value=None,
                            new_value=None,
                            username=session.get('full_name', 'User'),
                            notes=f'Bulk generation: {result["barcode"]}',
                            source_page='Barcode Management',
                            source_url=request.url
                        )
                    else:
                        failed_count += 1
                        failed_variants.append(f"{product_code}-{color_name}")
                else:
                    failed_count += 1
                    failed_variants.append(f"{product_code}-{color_name}")
                    
            except Exception as e:
                print(f"❌ Error generating barcode for {product_code}-{color_name}: {e}")
                failed_count += 1
                failed_variants.append(f"{product_code}-{color_name}")
        
        return jsonify({
            'success': True,
            'success_count': success_count,
            'failed_count': failed_count,
            'failed_variants': failed_variants[:10],  # Return first 10 failures
            'total': len(variants),
            'message': f'Generated {success_count} barcodes successfully'
        })
        
    except Exception as e:
        print(f"❌ Error in bulk generation: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/barcode/generate_selected', methods=['POST'])
@action_permission_required('barcode_system')
def generate_selected_barcodes():
    """Generate barcodes for selected variants"""
    try:
        data = request.get_json()
        variant_ids = data.get('variant_ids', [])
        
        if not variant_ids:
            return jsonify({
                'success': False,
                'error': 'No variants selected'
            }), 400
        
        from barcode_utils import generate_barcode_for_variant
        
        success_count = 0
        failed_count = 0
        failed_items = []
        user_id = session.get('user_id', 0)
        
        for variant_id in variant_ids:
            try:
                # Check if barcode already exists
                existing = db.get_barcode_by_variant(variant_id)
                if existing:
                    failed_count += 1
                    failed_items.append(f"Variant {variant_id}: Barcode already exists")
                    continue
                
                # Get variant details
                variant = db.get_variant_details_for_barcode(variant_id)
                
                if not variant:
                    failed_count += 1
                    failed_items.append(f"Variant {variant_id} not found")
                    continue
                
                product_code = variant[1]
                color_name = variant[4]
                
                # Generate barcode image
                result = generate_barcode_for_variant(product_code, color_name, variant_id)
                
                if not result:
                    failed_count += 1
                    failed_items.append(f"{product_code} - {color_name}: Failed to generate")
                    continue
                
                # ✅ حفظ في database
                barcode_id = db.create_barcode(
                    variant_id=variant_id,
                    barcode_number=result['barcode'],
                    image_path=result['image_path'],
                    user_id=user_id
                )
                
                if barcode_id:
                    success_count += 1
                    
                    # Log the operation
                    db.add_stock_log(
                        operation_type="Barcode Generated (Selected)",
                        product_id=None,
                        variant_id=variant_id,
                        product_code=product_code,
                        brand_name=variant[2],
                        product_type=variant[3],
                        color_name=color_name,
                        image_url=variant[7] or '',
                        old_value=None,
                        new_value=result['barcode'],
                        username=session.get('fullname', 'User'),
                        notes=f"Selected generation",
                        source_page="Barcode Management",
                        source_url=request.url
                    )
                else:
                    failed_count += 1
                    failed_items.append(f"{product_code} - {color_name}: Failed to save")
                    
            except Exception as e:
                failed_count += 1
                failed_items.append(f"Variant {variant_id}: {str(e)}")
                print(f"❌ Error generating barcode for variant {variant_id}: {e}")
        
        return jsonify({
            'success': True,
            'success_count': success_count,
            'failed_count': failed_count,
            'failed_items': failed_items[:10],
            'message': f'Generated {success_count} barcodes ({failed_count} failed)'
        })
        
    except Exception as e:
        print(f"❌ Error in selected generation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@app.route('/barcode/view/<int:variant_id>')
@action_permission_required('barcode_system')
def view_barcode(variant_id):
    """Get barcode details for a variant (API endpoint)"""
    try:
        # Get variant details
        variant = db.get_variant_details_for_barcode(variant_id)
        
        if not variant:
            return jsonify({
                'success': False,
                'error': 'Variant not found'
            }), 404
        
        # Get barcode details
        barcode = db.get_barcode_by_variant(variant_id)
        
        if not barcode:
            return jsonify({
                'success': False,
                'error': 'Barcode not found for this variant'
            }), 404
        
        # Prepare barcode image URL
        barcode_image_url = None
        if barcode[3]:  # image_path
            # Convert file path to URL
            if barcode[3].startswith('static/'):
                barcode_image_url = '/' + barcode[3]
            else:
                barcode_image_url = barcode[3]
        
        return jsonify({
            'success': True,
            'product_code': variant[1],
            'brand_name': variant[2],
            'product_type': variant[3],
            'color_name': variant[4],
            'current_stock': variant[6],
            'image_url': variant[7],
            'wholesale_price': variant[9] if len(variant) > 9 else 0,
            'retail_price': variant[10] if len(variant) > 10 else 0,
            'barcode': barcode[2],
            'barcode_image_url': barcode_image_url,
            'generated_at': str(barcode[4]) if barcode[4] else None
        })
        
    except Exception as e:
        print(f"❌ Error viewing barcode: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# === Barcode Scanner Page ===

@app.route('/barcode/scanner')
@page_permission_required('barcode_system')
def barcode_scanner():
    """Barcode scanner page"""
    user_id = session.get('user_id', 0)
    
    # Get active session
    active_session_db = db.get_active_session(user_id)
    
    if active_session_db:
        session_id = active_session_db[0]
        
        # ✅ جيب الـ items مع كل التفاصيل!
        items_with_details = db.get_session_items_with_details(session_id)
        
        # ✅ حوّل الـ image URLs
        for item in items_with_details:
            image_url = item.get('image_url')
            if image_url:
                if image_url.startswith('http://') or image_url.startswith('https://'):
                    item['image_url'] = image_url  # خليها زي ما هي
                else:
                    item['image_url'] = f"/{image_url}"  # حط / قدام static paths
        
        session_data = {
            'exists': True,
            'id': int(active_session_db[0]),
            'mode': str(active_session_db[2]),
            'items': items_with_details,  # ✅ بعت الـ items الكاملة
            'itemsjson': json.dumps(items_with_details),  # ✅ JSON string للـ JavaScript
            'createdat': str(active_session_db[4]) if len(active_session_db) > 4 else ''
        }
    else:
        session_data = {
            'exists': False,
            'id': 0,
            'mode': 'add',
            'items': [],
            'itemsjson': '[]',
            'createdat': ''
        }
    
    return render_template('barcode_scanner.html', active_session=session_data)

@app.route('/barcode/session/start', methods=['POST'])
@action_permission_required('barcode_system')
def start_scan_session():
    """Start a new scanning session"""
    try:
        data = request.get_json()
        mode = data.get('mode', 'add')  # 'add' or 'remove'
        
        if mode not in ['add', 'remove']:
            return jsonify({
                'success': False,
                'error': 'Invalid mode. Must be "add" or "remove"'
            }), 400
        
        user_id = session.get('user_id', 0)
        
        # Check if there's already an active session
        existing = db.get_active_session(user_id)
        if existing:
            return jsonify({
                'success': False,
                'error': 'You already have an active session. Please close it first.',
                'session_id': existing[0]
            }), 400
        
        # Create new session
        session_id = db.create_scan_session(user_id, mode)
        
        if not session_id:
            return jsonify({
                'success': False,
                'error': 'Failed to create session'
            }), 500
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'mode': mode,
            'message': f'Session started in {mode.upper()} mode'
        })
        
    except Exception as e:
        print(f"❌ Error starting session: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/barcode/session/scan', methods=['POST'])
@page_permission_required('barcode_system')
def barcode_session_scan():
    """Scan barcode and add to session"""
    try:
        data = request.get_json()
        barcode = str(data.get('barcode', '')).strip()
        
        # ✅ حافظ على الأصفار للـ EAN-13
        if barcode.isdigit() and len(barcode) <= 13:
            barcode = barcode.zfill(13)  # يضيف أصفار لو ناقصة
        
        print(f"📊 Scanning barcode: {barcode}")
        user_id = session.get('user_id', 0)
        
        if not barcode:
            return jsonify({'success': False, 'error': 'Barcode is required'})
        
        # Get active session
        active_session = db.get_active_session(user_id)
        if not active_session:
            return jsonify({'success': False, 'error': 'No active session. Please start a session first.'})
        
        session_id = active_session[0]
        
        # Get variant by barcode number
        variant = db.get_variant_by_barcode(barcode)
        if not variant:
            return jsonify({'success': False, 'error': f'Product not found with barcode: {barcode}'})
        
        # Parse variant data
        variant_id = variant[0]
        product_code = variant[1]
        brand_name = variant[2]
        product_type = variant[3]
        color_name = variant[4]
        color_code = variant[5]
        stock_quantity = variant[6]
        image_url = variant[7]
        product_size = variant[8] if len(variant) > 8 else None
        
        # Add item to session
        result = db.add_item_to_session(session_id, variant_id)
        
        if result:
            # ✅ إصلاح الـ image URL - نتأكد لو بيبدأ بـ http مش نحط /
            if image_url:
                if image_url.startswith('http://') or image_url.startswith('https://'):
                    final_image_url = image_url  # ✅ خليها زي ما هي
                else:
                    final_image_url = f"/{image_url}"  # ✅ حط / قدام static paths
            else:
                final_image_url = None
            
            # Return full item data - استخدم نفس الأسماء اللي في JavaScript!
            item_data = {
                'variant_id': variant_id,
                'productcode': product_code,      # ✅ بدون underscore
                'brandname': brand_name,          # ✅ بدون underscore
                'producttype': product_type,      # ✅ بدون underscore
                'colorname': color_name,          # ✅ بدون underscore
                'colorcode': color_code,          # ✅ بدون underscore
                'stockquantity': stock_quantity,  # ✅ بدون underscore
                'imageurl': final_image_url,      # ✅ بدون underscore
                'productsize': product_size,      # ✅ بدون underscore
                'quantity': 1
            }
            
            return jsonify({
                'success': True,
                'message': f'✅ Added: {product_code} - {color_name}',
                'item': item_data
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to add item to session'})
            
    except Exception as e:
        print(f"❌ Error in barcode_session_scan: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'})


@app.route('/barcode/session/update', methods=['POST'])
@action_permission_required('barcode_system')
def update_session_item():
    """Update quantity for an item in session OR update all items"""
    try:
        data = request.get_json()
        user_id = session.get('user_id', 0)
        
        active_session = db.get_active_session(user_id)
        if not active_session:
            return jsonify({'success': False, 'error': 'No active session'}, 400)
        
        session_id = active_session[0]
        items = json.loads(active_session[3]) if active_session[3] else []
        
        # ✅ Check if updating all items or single item
        if 'items' in data:
            # Update all items at once (from quantity modal)
            new_items = []
            for item in data['items']:
                new_items.append({
                    'variant_id': item.get('variant_id'),
                    'quantity': item.get('quantity', 1)
                })
            items = new_items
            
        elif 'variant_id' in data:
            # Update single item quantity (original functionality)
            variant_id = data.get('variant_id')
            quantity = data.get('quantity', 1)
            
            if not variant_id:
                return jsonify({'success': False, 'error': 'Variant ID is required'}, 400)
            
            # Find and update item
            item = next((item for item in items if item['variant_id'] == variant_id), None)
            if not item:
                return jsonify({'success': False, 'error': 'Item not found in session'}, 404)
            
            item['quantity'] = max(1, int(quantity))  # Minimum 1
        else:
            return jsonify({'success': False, 'error': 'Invalid request'}, 400)
        
        # Update session
        success = db.update_session_items(session_id, json.dumps(items))
        
        if not success:
            return jsonify({'success': False, 'error': 'Failed to update session'}, 500)
        
        return jsonify({
            'success': True,
            'message': 'Session updated',
            'session': {
                'total_items': len(items),
                'total_quantity': sum(item['quantity'] for item in items)
            }
        })
        
    except Exception as e:
        print(f"❌ Error updating item: {e}")
        return jsonify({'success': False, 'error': str(e)}, 500)


@app.route('/barcode/session/remove', methods=['POST'])
@action_permission_required('barcode_system')
def remove_session_item():
    """Remove an item from session"""
    try:
        data = request.get_json()
        variant_id = data.get('variant_id')
        
        if not variant_id:
            return jsonify({
                'success': False,
                'error': 'Variant ID is required'
            }), 400
        
        user_id = session.get('user_id', 0)
        active_session = db.get_active_session(user_id)
        
        if not active_session:
            return jsonify({
                'success': False,
                'error': 'No active session'
            }), 400
        
        session_id = active_session[0]
        items = json.loads(active_session[3]) if active_session[3] else []
        
        # Remove item
        items = [item for item in items if item['variant_id'] != variant_id]
        
        # Update session
        success = db.update_session_items(session_id, json.dumps(items))
        
        if not success:
            return jsonify({
                'success': False,
                'error': 'Failed to update session'
            }), 500
        
        return jsonify({
            'success': True,
            'message': 'Item removed from session',
            'session': {
                'total_items': len(items),
                'total_quantity': sum(item['quantity'] for item in items)
            }
        })
        
    except Exception as e:
        print(f"❌ Error removing item: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/barcode/session/clear', methods=['POST'])
@action_permission_required('barcode_system')
def clear_session():
    """Clear all items from session"""
    try:
        user_id = session.get('user_id', 0)
        active_session = db.get_active_session(user_id)
        
        if not active_session:
            return jsonify({
                'success': False,
                'error': 'No active session'
            }), 400
        
        session_id = active_session[0]
        
        # Clear items
        success = db.update_session_items(session_id, '[]')
        
        if not success:
            return jsonify({
                'success': False,
                'error': 'Failed to clear session'
            }), 500
        
        return jsonify({
            'success': True,
            'message': 'Session cleared'
        })
        
    except Exception as e:
        print(f"❌ Error clearing session: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/barcode/session/confirm', methods=['POST'])
@action_permission_required('barcode_system')
def confirm_session():
    """Confirm session and update stock"""
    try:
        user_id = session.get('user_id', 0)
        
        # Get active session
        active_session = db.get_active_session(user_id)
        if not active_session:
            return jsonify({'success': False, 'error': 'No active session'})
        
        session_id = active_session[0]
        session_mode = active_session[2]
        
        # Get session items with full details
        items = db.get_session_items_with_details(session_id)
        if not items:
            return jsonify({'success': False, 'error': 'Session is empty'})
        
        # Prepare stock updates list
        stock_updates = []
        
        # Get connection
        conn = db.get_connection()
        cursor = conn.cursor()
        
        try:
            # Step 1: Get all details BEFORE updating
            for item in items:
                variant_id = item['variant_id']
                quantity = item['quantity']
                
                try:
                    # Get current details from database
                    cursor.execute("""
                        SELECT pv.current_stock, bp.id, bp.product_code, 
                               br.brand_name, pt.type_name, c.color_name, ci.image_url
                        FROM product_variants pv
                        JOIN base_products bp ON pv.base_product_id = bp.id
                        JOIN brands br ON bp.brand_id = br.id
                        JOIN product_types pt ON bp.product_type_id = pt.id
                        JOIN colors c ON pv.color_id = c.id
                        LEFT JOIN color_images ci ON pv.id = ci.variant_id
                        WHERE pv.id = ?
                    """, (variant_id,))
                    
                    result = cursor.fetchone()
                    
                    if result:
                        old_stock = result[0]
                        
                        # Calculate new stock
                        if session_mode == 'add':
                            new_stock = old_stock + quantity
                        else:  # remove
                            new_stock = max(0, old_stock - quantity)
                        
                        # Only add to updates if stock actually changed
                        if old_stock != new_stock:
                            stock_updates.append({
                                'variant_id': variant_id,
                                'product_id': result[1],
                                'product_code': result[2],
                                'brand_name': result[3],
                                'product_type': result[4],
                                'color_name': result[5],
                                'image_url': result[6] or '',
                                'old_stock': old_stock,
                                'new_stock': new_stock,
                                'quantity': quantity,
                                'skip_log': False
                            })
                        else:
                            stock_updates.append({'skip_log': True})
                    
                except Exception as e:
                    print(f"❌ Error getting details for variant {variant_id}: {e}")
                    stock_updates.append({'skip_log': True})
            
            # Step 2: Update stock for all items
            for update in stock_updates:
                if not update.get('skip_log', False):
                    cursor.execute("""
                        UPDATE product_variants 
                        SET current_stock = ? 
                        WHERE id = ?
                    """, (update['new_stock'], update['variant_id']))
            
            # Commit all updates
            conn.commit()
            conn.close()
            
            # Step 3: Log successful updates (after commit)
            logged_count = 0
            for update in stock_updates:
                if not update.get('skip_log', False) and 'old_stock' in update:
                    db.add_stock_log(
                        operation_type=f'Barcode Scan - {session_mode.title()}',
                        product_id=update['product_id'],
                        variant_id=update['variant_id'],
                        product_code=update['product_code'],
                        brand_name=update['brand_name'],
                        product_type=update['product_type'],
                        color_name=update['color_name'],
                        image_url=update['image_url'],
                        old_value=update['old_stock'],
                        new_value=update['new_stock'],
                        username=session.get('full_name', 'User'),
                        notes=f"Stock {'increased' if session_mode == 'add' else 'decreased'} by {update['quantity']} units via barcode scanner",
                        source_page='Barcode Scanner',
                        source_url=request.url
                    )
                    logged_count += 1
            
            # Close session
            db.close_session(session_id, 'confirmed')
            
            return jsonify({
                'success': True, 
                'updated_count': logged_count,
                'message': f'Stock updated for {logged_count} items'
            })
            
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'success': False, 'error': f'Failed to update stock: {str(e)}'})
    
    except Exception as e:
        print(f"❌ Error confirming session: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/barcode/session/cancel', methods=['POST'])
@action_permission_required('barcode_system')
def cancel_session():
    """Cancel active session without updating stock"""
    try:
        user_id = session.get('user_id', 0)
        active_session = db.get_active_session(user_id)
        
        if not active_session:
            return jsonify({
                'success': False,
                'error': 'No active session'
            }), 400
        
        session_id = active_session[0]
        
        # Close session
        success = db.close_session(session_id, 'cancelled')
        
        if not success:
            return jsonify({
                'success': False,
                'error': 'Failed to cancel session'
            }), 500
        
        return jsonify({
            'success': True,
            'message': 'Session cancelled'
        })
        
    except Exception as e:
        print(f"❌ Error cancelling session: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# === Barcode Printing Page ===

@app.route('/barcode/printing')
@page_permission_required('barcode_system')
def barcode_printing():
    """Barcode printing page"""
    # Get filters
    search = request.args.get('search', '')
    brand_filter = request.args.get('brand', '')
    type_filter = request.args.get('type', '')
    color_filter = request.args.get('color', '')
    stock_only = request.args.get('stock_only', '') == 'on'
    page = int(request.args.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page
    
    # Get filter options
    brands = db.get_brands_for_filter()
    types = [pt[1] for pt in db.get_all_product_types()]
    colors = [c[1] for c in db.get_all_colors()]
    
    # Get variants with barcodes
    variants = db.get_variants_with_barcode(
        search=search,
        brand_filter=brand_filter,
        type_filter=type_filter,
        color_filter=color_filter,
        limit=per_page,
        offset=offset
    )
    
    # Filter by stock if needed
    if stock_only:
        variants = [v for v in variants if v[6] > 0]  # current_stock > 0
    
    total_count = db.count_variants_with_barcode(
        search=search,
        brand_filter=brand_filter,
        type_filter=type_filter,
        color_filter=color_filter
    )
    
    # Calculate pagination
    total_pages = (total_count + per_page - 1) // per_page
    has_more = page < total_pages
    
    return render_template('barcode_printing.html',
                         variants=variants,
                         brands=brands,
                         types=types,
                         colors=colors,
                         search=search,
                         brand_filter=brand_filter,
                         type_filter=type_filter,
                         color_filter=color_filter,
                         stock_only=stock_only,
                         page=page,
                         total_pages=total_pages,
                         total_count=total_count,
                         has_more=has_more)


@app.route('/barcode/print', methods=['POST'])
@action_permission_required('barcode_system')
def print_barcodes():
    """Generate PDF with barcode labels"""
    try:
        data = request.get_json()
        variant_ids = data.get('variant_ids', [])
        quantities = data.get('quantities', {})  # {variant_id: quantity}
        
        if not variant_ids:
            return jsonify({
                'success': False,
                'error': 'No variants selected'
            }), 400
        
        # Prepare labels data
        labels_data = []
        
        for variant_id in variant_ids:
            # Get barcode info
            barcode = db.get_barcode_by_variant(variant_id)
            
            if not barcode:
                continue
            
            # Get variant details
            variant = db.get_variant_details_for_barcode(variant_id)
            
            if not variant:
                continue
            
            # Get quantity
            quantity = quantities.get(str(variant_id), 1)
            
            labels_data.append({
                'barcode_image': barcode[3],  # image_path
                'product_code': variant[1],
                'color_name': variant[4],
                'barcode_number': barcode[2],  # barcode_number
                'quantity': int(quantity)
            })
        
        if not labels_data:
            return jsonify({
                'success': False,
                'error': 'No valid barcodes found'
            }), 400
        
        # Generate PDF
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f'barcode_labels_{timestamp}.pdf'
        pdf_path = os.path.join('static', 'temp', pdf_filename)
        
        # Create temp directory if it doesn't exist
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
        
        # Create PDF
        result = create_barcode_labels_pdf(labels_data, pdf_path)
        
        if not result:
            return jsonify({
                'success': False,
                'error': 'Failed to generate PDF'
            }), 500
        
        # Log each product separately with full details
        logged_count = 0
        conn = db.get_connection()
        cursor = conn.cursor()
        
        try:
            for variant_id in variant_ids:
                quantity = quantities.get(str(variant_id), 1)
                
                try:
                    # Get full product details from database
                    cursor.execute("""
                        SELECT bp.id, bp.product_code, br.brand_name, 
                               pt.type_name, c.color_name, pv.current_stock, ci.image_url
                        FROM product_variants pv
                        JOIN base_products bp ON pv.base_product_id = bp.id
                        JOIN brands br ON bp.brand_id = br.id
                        JOIN product_types pt ON bp.product_type_id = pt.id
                        JOIN colors c ON pv.color_id = c.id
                        LEFT JOIN color_images ci ON pv.id = ci.variant_id
                        WHERE pv.id = ?
                    """, (variant_id,))
                    
                    result = cursor.fetchone()
                    
                    if result:
                        db.add_stock_log(
                            operation_type='Barcode Labels Printed',
                            product_id=result[0],
                            variant_id=variant_id,
                            product_code=result[1],
                            brand_name=result[2],
                            product_type=result[3],
                            color_name=result[4],
                            image_url=result[6] or '',
                            old_value=result[5],
                            new_value=result[5],
                            username=session.get('full_name', 'User'),
                            notes=f'Printed {quantity} barcode label{"s" if quantity > 1 else ""}',
                            source_page='Barcode Printing',
                            source_url=request.url
                        )
                        logged_count += 1
                
                except Exception as e:
                    print(f"❌ Error logging barcode print for variant {variant_id}: {e}")
                    continue
            
            conn.close()
            print(f"✅ Logged {logged_count} barcode print operations")
            
        except Exception as e:
            print(f"❌ Error in barcode print logging: {e}")
            if conn:
                conn.close()
        
        # Calculate total labels (بعد الـ Logging)
        total_labels = sum(item['quantity'] for item in labels_data)
        
        return jsonify({
            'success': True,
            'pdf_url': f'/static/temp/{pdf_filename}',
            'total_labels': total_labels,
            'total_products': len(labels_data),
            'message': f'PDF generated with {total_labels} labels'
        })

        
    except Exception as e:
        print(f"❌ Error printing barcodes: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# === Barcode Lookup API ===

@app.route('/barcode/lookup')
@login_required
def barcode_lookup_page():
    """Barcode Lookup Page - Quick scan without stock changes"""
    return render_template('barcode_lookup.html')

# الـ API موجود أصلاً - بس نتأكد إنه شغال
@app.route('/api/barcode/lookup/<barcode>')
@login_required
def barcode_lookup(barcode):
    """Quick barcode lookup API"""
    try:
        # ✅ حافظ على الأصفار
        barcode = str(barcode).strip()
        if barcode.isdigit() and len(barcode) <= 13:
            barcode = barcode.zfill(13)
        
        print(f"🔍 Looking up barcode: {barcode}")
        
        barcode_data = db.get_barcode_by_number(barcode)

        
        if not barcode_data:
            return jsonify({'success': False, 'error': 'Barcode not found'})
        
        # Fix image URL
        image_url = barcode_data[10] or None
        if image_url:
            if image_url.startswith('http://') or image_url.startswith('https://'):
                final_image_url = image_url
            else:
                final_image_url = f"/{image_url}"
        else:
            final_image_url = None
        
        return jsonify({
            'success': True,
            'variant_id': barcode_data[1],
            'barcode': barcode_data[2],
            'product_code': barcode_data[5],
            'brand_name': barcode_data[6],
            'product_type': barcode_data[7],
            'color_name': barcode_data[8],
            'color_code': barcode_data[9],
            'current_stock': barcode_data[4],
            'wholesale_price': barcode_data[11],
            'retail_price': barcode_data[12],
            'product_size': barcode_data[13],
            'image_url': final_image_url
        })
        
    except Exception as e:
        print(f"❌ Error looking up barcode: {e}")
        return jsonify({'success': False, 'error': str(e)})


# === Cleanup Task (Optional - run periodically) ===

@app.route('/admin/cleanup_sessions')
@page_permission_required('barcode_system')
def cleanup_sessions():
    """Manual cleanup of old sessions (can be automated)"""
    try:
        count = db.cleanup_old_sessions(hours=24)
        flash(f'Cleaned up {count} old sessions', 'success')
    except Exception as e:
        flash(f'Error cleaning up sessions: {str(e)}', 'error')
    
    return redirect(url_for('barcode_scanner'))


# ========================================
# BARCODE IMAGE REGENERATION
# ========================================

@app.route('/barcode/regenerate/<int:variant_id>', methods=['POST'])
@action_permission_required('barcode_system')
def regenerate_single_barcode_image(variant_id):
    """Regenerate barcode image for a single variant"""
    try:
        # Get variant details
        variant = db.get_variant_details_for_barcode(variant_id)
        
        if not variant:
            return jsonify({
                'success': False,
                'error': 'Variant not found'
            }), 404
        
        product_code = variant[1]
        color_name = variant[4]
        
        # Import barcode utils
        from barcode_utils import generate_barcode_for_variant
        
        # Generate barcode image
        result = generate_barcode_for_variant(product_code, color_name, variant_id)
        
        if not result:
            return jsonify({
                'success': False,
                'error': 'Failed to generate barcode image'
            }), 500
        
        # Update database
        db.update_barcode_image_path(variant_id, result['image_path'])
        
        # Log the operation
        db.add_stock_log(
            operation_type="Barcode Image Regenerated",
            product_id=None,
            variant_id=variant_id,
            product_code=product_code,
            brand_name=variant[2],
            product_type=variant[3],
            color_name=color_name,
            image_url=variant[7] or '',
            old_value=None,
            new_value=None,
            username=session.get('fullname', 'User'),
            notes=f"Barcode image regenerated manually",
            source_page="Barcode Management",
            source_url=request.url
        )
        
        return jsonify({
            'success': True,
            'message': f'✅ Barcode image regenerated for {product_code} - {color_name}',
            'image_path': result['image_path']
        })
        
    except Exception as e:
        print(f"❌ Error regenerating barcode image: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/barcode/regenerate_bulk', methods=['POST'])
@action_permission_required('barcode_system')
def regenerate_bulk_barcode_images():
    """Regenerate barcode images for multiple variants"""
    try:
        data = request.get_json()
        variant_ids = data.get('variant_ids', [])
        
        if not variant_ids:
            return jsonify({
                'success': False,
                'error': 'No variants selected'
            }), 400
        
        from barcode_utils import generate_barcode_for_variant
        
        success_count = 0
        failed_count = 0
        failed_items = []
        
        for variant_id in variant_ids:
            try:
                # Get variant details
                variant = db.get_variant_details_for_barcode(variant_id)
                
                if not variant:
                    failed_count += 1
                    failed_items.append(f"Variant {variant_id} not found")
                    continue
                
                product_code = variant[1]
                color_name = variant[4]
                
                # Generate barcode image
                result = generate_barcode_for_variant(product_code, color_name, variant_id)
                
                if result:
                    # Update database
                    db.update_barcode_image_path(variant_id, result['image_path'])
                    success_count += 1
                    
                    # Log
                    db.add_stock_log(
                        operation_type="Barcode Image Regenerated (Bulk)",
                        product_id=None,
                        variant_id=variant_id,
                        product_code=product_code,
                        brand_name=variant[2],
                        product_type=variant[3],
                        color_name=color_name,
                        image_url=variant[7] or '',
                        old_value=None,
                        new_value=None,
                        username=session.get('fullname', 'User'),
                        notes=f"Bulk regeneration",
                        source_page="Barcode Management",
                        source_url=request.url
                    )
                else:
                    failed_count += 1
                    failed_items.append(f"{product_code} - {color_name}")
                    
            except Exception as e:
                failed_count += 1
                failed_items.append(f"Variant {variant_id}: {str(e)}")
        
        return jsonify({
            'success': True,
            'success_count': success_count,
            'failed_count': failed_count,
            'failed_items': failed_items[:10],  # First 10 failures
            'message': f'✅ Regenerated {success_count} images ({failed_count} failed)'
        })
        
    except Exception as e:
        print(f"❌ Error in bulk regeneration: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

# إضافة health check endpoint
@app.route('/health')
def health_check():
    return {'status': 'healthy', 'timestamp': datetime.now().isoformat()}




if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    print(f"🚀 Starting on Render...")
    print(f"📊 Port: {port}")
    print(f"🔧 Debug: {debug}")
    
    app.run(debug=debug, host='0.0.0.0', port=port)
